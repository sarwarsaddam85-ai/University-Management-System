from flask import Blueprint, render_template, request, redirect, url_for, session, flash, make_response, current_app
from app import db
from app.decorators import admin_required, login_required
import datetime
from werkzeug.utils import secure_filename
import os
# --- Imports for Excel Export ---
from openpyxl import Workbook
import openpyxl.styles # For cell formatting
import io # For saving Excel in memory
# --- FPDF imports are removed ---

# Define the blueprint
finance_bp = Blueprint(
    'finance',
    __name__,
    template_folder='../templates/finance'
)

# --- PDF CHALLAN CLASS DEFINITION is REMOVED ---

# --- ADMIN: VIEW FEE STATUS PAGE ---
@finance_bp.route('/status')
@admin_required
def fee_status():
    students_list = []
    try:
        students_ref = db.collection('students').stream()
        for student in students_ref:
            student_data = student.to_dict()
            student_data['id'] = student.id # User ID
            student_data['fee_assigned'] = student_data.get('fee_assigned', 0.0)
            student_data['fee_paid'] = student_data.get('fee_paid', 0.0)
            student_data['fee_status'] = student_data.get('fee_status', 'Pending Assignment')
            student_data['receipt_url'] = student_data.get('receipt_url', None)
            students_list.append(student_data)
    except Exception as e:
        flash(f'Error fetching student fee status: {e}', 'error')
    return render_template('fee_status.html', students=students_list)


# --- ADMIN: ASSIGN/UPDATE FEE PAGE (Includes Notification) ---
@finance_bp.route('/assign/<string:student_id>', methods=['GET', 'POST'])
@admin_required
def assign_fee(student_id):
    student_ref = db.collection('students').document(student_id)
    if request.method == 'POST':
        try:
            assigned_fee = float(request.form.get('assigned_fee', 0.0))
            paid_fee = float(request.form.get('paid_fee', 0.0))
            status = 'Pending Assignment'
            if assigned_fee > 0:
                if paid_fee >= assigned_fee: status = 'Paid'
                elif paid_fee > 0: status = 'Partially Paid'
                else: status = 'Unpaid'

            fee_update_data = {
                'fee_assigned': assigned_fee,
                'fee_paid': paid_fee,
                'fee_status': status,
                'fee_updated_at': datetime.datetime.utcnow()
            }
            student_ref.update(fee_update_data)

            # --- Create Notification ---
            try:
                notification_message = f"Your fee information has been updated. Status: {status}, Assigned: ${assigned_fee:,.2f}, Paid: ${paid_fee:,.2f}"
                notification_data = {
                    'student_id': student_id,
                    'message': notification_message,
                    'type': 'fee',
                    'link': url_for('finance.my_fees', _external=True),
                    'created_at': datetime.datetime.utcnow(),
                    'is_read': False
                }
                db.collection('notifications').add(notification_data)
            except Exception as notify_error:
                print(f"--- FAILED to create fee notification for {student_id}: {notify_error} ---")
                flash(f'Fee updated, but failed to send notification: {notify_error}', 'warning')
            # --- End Notification ---

            flash('Fee information updated successfully!', 'success')
            return redirect(url_for('finance.fee_status'))

        except ValueError:
            flash('Invalid input. Please enter numbers for fee amounts.', 'error')
        except Exception as e:
            flash(f'Error updating fee information: {e}', 'error')

    # --- GET Request ---
    try:
        student = student_ref.get()
        if not student.exists:
            flash('Student not found.', 'error')
            return redirect(url_for('finance.fee_status'))
        student_data = student.to_dict()
        student_data.setdefault('fee_assigned', 0.0)
        student_data.setdefault('fee_paid', 0.0)
    except Exception as e:
        flash(f'Error fetching student data: {e}', 'error')
        return redirect(url_for('finance.fee_status'))
    return render_template('assign_fee.html', student=student_data, student_id=student_id)


# --- STUDENT: VIEW MY FEES PAGE ---
@finance_bp.route('/my_fees')
@login_required
def my_fees():
    student_id = session.get('user_id')
    student_fee_data = None
    try:
        student_ref = db.collection('students').document(student_id)
        student = student_ref.get()
        if student.exists:
            student_data = student.to_dict()
            student_fee_data = {
                'fee_assigned': student_data.get('fee_assigned', 0.0),
                'fee_paid': student_data.get('fee_paid', 0.0),
                'fee_status': student_data.get('fee_status', 'Not Assigned'),
                'receipt_url': student_data.get('receipt_url', None)
            }
    except Exception as e:
        flash(f'Error fetching your fee status: {e}', 'error')
    return render_template('my_fees.html', student_fee_data=student_fee_data)


# --- STUDENT: UPLOAD RECEIPT ---
@finance_bp.route('/upload_receipt', methods=['POST'])
@login_required
def upload_receipt():
    student_id = session.get('user_id')
    student_ref = db.collection('students').document(student_id)
    try:
        receipt_file = request.files['receipt_file']
        if not receipt_file:
            flash('No receipt file selected.', 'error')
            return redirect(url_for('finance.my_fees'))
        filename = secure_filename(receipt_file.filename)
        unique_filename = f"{student_id}_{filename}"
        receipts_dir = os.path.join(current_app.root_path, 'static/receipts')
        file_path = os.path.join(receipts_dir, unique_filename)
        os.makedirs(receipts_dir, exist_ok=True)
        receipt_file.save(file_path)
        file_url = url_for('static', filename=f'receipts/{unique_filename}')
        student_ref.update({
            'receipt_url': file_url,
            'receipt_uploaded_at': datetime.datetime.utcnow()
        })
        flash('Receipt uploaded successfully!', 'success')
    except Exception as e:
        flash(f'Error uploading receipt: {e}', 'error')
    return redirect(url_for('finance.my_fees'))


# --- ADMIN: VIEW HTML CHALLAN ---
@finance_bp.route('/view_challan/<string:student_id>')
@admin_required
def view_challan(student_id):
    try:
        student_ref = db.collection('students').document(student_id)
        student = student_ref.get()
        if not student.exists:
            flash('Student not found.', 'error')
            return redirect(url_for('finance.fee_status'))
        student_data = student.to_dict()
        student_data['id'] = student.id

        today_date = datetime.date.today()
        due_date = today_date + datetime.timedelta(days=15)

        return render_template('challan_template.html',
                               student=student_data,
                               today_date=today_date,
                               due_date=due_date)
    except Exception as e:
        flash(f'Error generating challan page: {e}', 'error')
        return redirect(url_for('finance.fee_status'))


# --- ADMIN: EXPORT FEE STATUS TO EXCEL ---
@finance_bp.route('/export/excel')
@admin_required
def export_fees_excel():
    try:
        # 1. Fetch data
        students_list = []
        students_ref = db.collection('students').stream()
        for student in students_ref:
            student_data = student.to_dict()
            student_data['id'] = student.id
            student_data['fee_assigned'] = student_data.get('fee_assigned', 0.0)
            student_data['fee_paid'] = student_data.get('fee_paid', 0.0)
            student_data['fee_status'] = student_data.get('fee_status', 'Pending Assignment')
            student_data['balance'] = student_data['fee_assigned'] - student_data['fee_paid']
            students_list.append(student_data)

        # 2. Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Fee Status Report"
        headers = ["Student Name", "Roll No.", "Department", "Semester", "Assigned Fee", "Amount Paid", "Remaining Balance", "Status"]
        ws.append(headers)
        for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Write Data
        for student in students_list:
            row = [
                student.get('name', 'N/A'),
                student.get('roll_no', 'N/A'),
                student.get('department', 'N/A'),
                student.get('semester', 'N/A'),
                student.get('fee_assigned', 0.0),
                student.get('fee_paid', 0.0),
                student.get('balance', 0.0),
                student.get('fee_status', 'N/A')
            ]
            ws.append(row)
            ws[f'E{ws.max_row}'].number_format = '"$"#,##0.00'
            ws[f'F{ws.max_row}'].number_format = '"$"#,##0.00'
            ws[f'G{ws.max_row}'].number_format = '"$"#,##0.00'

        # Adjust widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # 3. Save to memory buffer
        virtual_workbook = io.BytesIO()
        wb.save(virtual_workbook)
        excel_data = virtual_workbook.getvalue()

        # 4. Create Response
        response = make_response(excel_data)
        filename = f"Fee_Status_Report_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

    except Exception as e:
        flash(f'Error generating Excel report: {e}', 'error')
        return redirect(url_for('finance.fee_status'))
    
# --- STUDENT: SHOW FAKE PAYMENT PAGE (GET) ---
@finance_bp.route('/pay')
@login_required
def show_fake_payment():
    student_id = session.get('user_id')
    amount_due = 0.0
    try:
        student_ref = db.collection('students').document(student_id)
        student = student_ref.get()
        if student.exists:
            student_data = student.to_dict()
            fee_assigned = student_data.get('fee_assigned', 0.0)
            fee_paid = student_data.get('fee_paid', 0.0)
            amount_due = fee_assigned - fee_paid
            if amount_due <= 0:
                flash('You have no outstanding balance to pay.', 'info')
                return redirect(url_for('finance.my_fees'))
        else:
            flash('Student profile not found.', 'error')
            return redirect(url_for('dashboard.student_dashboard'))
    
    except Exception as e:
        flash(f'Error fetching fee data: {e}', 'error')
        return redirect(url_for('finance.my_fees'))
    
    return render_template('fake_payment.html', amount_due=amount_due)


# --- STUDENT: PROCESS FAKE PAYMENT (POST) ---
@finance_bp.route('/process_payment', methods=['POST'])
@login_required
def process_fake_payment():
    student_id = session.get('user_id')
    student_ref = db.collection('students').document(student_id)
    
    try:
        # Get the total assigned fee to mark as fully paid
        student_doc = student_ref.get()
        if not student_doc.exists:
            flash('Student profile not found.', 'error')
            return redirect(url_for('finance.my_fees'))
            
        student_data = student_doc.to_dict()
        fee_assigned = student_data.get('fee_assigned', 0.0)
        amount_paid_str = request.form.get('amount_paid', '0.0') # Get amount from form
        
        # Update the student's document to "Paid"
        student_ref.update({
            'fee_paid': fee_assigned, # Mark as fully paid
            'fee_status': 'Paid',
            'fee_updated_at': datetime.datetime.utcnow()
        })
        
        # --- Create Notification ---
        try:
            notification_message = f"Your online payment of ${float(amount_paid_str):,.2f} was successful. Your status is now 'Paid'."
            notification_data = {
                'student_id': student_id,
                'message': notification_message,
                'type': 'fee',
                'link': url_for('finance.my_fees', _external=True),
                'created_at': datetime.datetime.utcnow(),
                'is_read': False
            }
            db.collection('notifications').add(notification_data)
        except Exception as notify_error:
            print(f"--- FAILED to create payment notification: {notify_error} ---")
            # Don't fail the whole payment if notification fails
        
        # Redirect to success page
        return redirect(url_for('finance.payment_success'))
        
    except Exception as e:
        flash(f'An error occurred while processing your payment: {e}', 'error')
        return redirect(url_for('finance.my_fees'))